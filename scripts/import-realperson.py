"""실제 사람의 카드 명세서를 제공자에 넣는다. 저장소에는 로직만, 데이터는 저장소 밖.

  실행:
    python3 scripts/import-realperson.py _archive/명세서.xlsx --name 홍길동 \\
        --social7 9001011 --phone 010-1234-5678 --dry-run
    (확인한 뒤 --dry-run 을 떼고 다시)

  환경: BASE (기본 http://localhost:8082)

## 왜 스크립트인가

제공자(8082)는 운영에서 외부에 발행되지 않는다(W7-2 격리). 화면을 붙이려면 본체에 프록시를
뚫어야 하는데, **한 번 하는 일을 위해 격리에 영구적인 구멍**을 내는 셈이다. 게다가 입력이
실명·주민앞7·전화번호라, 상시 열린 입력창을 만들 이유가 없다.

## 명세서 형식은 카드사가 정한다

실물은 머리말 몇 줄 뒤에 머리글이 오고, 칸이 열 개쯤 된다.

    과거 카드이용내역
    고객명 : 이*원
    요청기간 : 2026.01.01 ~ 2026.12.31
    거래일 | 확정일 | 이용카드(뒤4자리) | 상품구분 | 가맹점명 | 사업자번호 | 이용금액 | …

그래서 **머리글을 이름으로 찾는다.** 위치로 읽으면 사용자가 엑셀에서 칸을 지우고 순서를
맞춰야 하는데, 거기가 바로 오류가 주입되는 자리다. 파일을 그대로 넣을 수 있어야 한다.

적재 API 는 `날짜,가맹점,금액,업종코드,사업자번호` 5칸을 받는다. 형식 해석은 여기서 끝내고
API 계약은 단순하게 둔다.

## 사업자번호가 사전의 키다

번호가 없으면 확정 분류 사전이 아무리 차 있어도 안 붙는다. 그래서 넣기 전에
**씨앗과 몇 건이 겹치는지 먼저 세어 둔다** — 적재 뒤 `category2_source='DICT'` 가 그 수와
맞아야 사슬이 이어진 것이다. 안 맞으면 어딘가 끊긴 것이고, 숫자가 없으면 알 수가 없다.
"""
import argparse
import io
import json
import os
import re
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..'))
SEED = os.path.join(ROOT, '_archive', 'merchant-seed.tsv')

# 카드사마다 칸 이름이 다르다. 우선순위 순 — '이용금액'을 '금액'보다 앞에 둬야
# '결제금액(청구액)' 같은 칸이 섞인 명세서에서 엉뚱한 칸을 집지 않는다.
NAMES = {
    'date':     ['거래일', '이용일자', '거래일자', '승인일자', '매출일자', '이용일', '사용일자', '날짜'],
    'merchant': ['가맹점명', '이용하신곳', '이용가맹점', '가맹점', '사용처', '상호'],
    'amount':   ['이용금액', '승인금액', '거래금액', '사용금액', '결제금액', '금액'],
    'biz':      ['사업자번호', '사업자등록번호', '사업자'],
    'industry': ['업종코드', '업종'],
}


def norm(s):
    """칸 이름 비교용 — BOM·공백·괄호를 지운다. 엑셀 저장본은 BOM 으로 시작하기도 한다."""
    return re.sub(r'[\s()（）﻿]', '', str(s or ''))


def read_rows(path):
    """xlsx 든 csv 든 '셀의 2차원 목록'으로 돌려준다."""
    if path.lower().endswith(('.xlsx', '.xlsm')):
        try:
            import openpyxl
        except ImportError:
            sys.exit('xlsx 를 읽으려면 openpyxl 이 필요하다:  .venv/bin/pip install openpyxl')
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
    # CSV — 카드사 파일은 CP949 인 경우가 흔하다. UTF-8 부터 시도한다.
    for enc in ('utf-8-sig', 'cp949', 'euc-kr'):
        try:
            text = io.open(path, encoding=enc).read()
            break
        except UnicodeDecodeError:
            continue
    else:
        sys.exit(f'인코딩을 못 읽었다: {path}')
    import csv
    return [r for r in csv.reader(io.StringIO(text))]


def find_header(rows):
    """머리글 줄과 칸 배치를 찾는다. 못 찾으면 (None, None)."""
    for i, row in enumerate(rows[:30]):          # 머리말이 길어야 몇 줄이다
        cells = [norm(c) for c in row]
        got = {}
        for key, names in NAMES.items():
            for want in names:
                if norm(want) in cells:
                    got[key] = cells.index(norm(want))
                    break
        if all(k in got for k in ('date', 'merchant', 'amount')):
            return i, got
    return None, None


def to_csv(rows, header_at, cols):
    """적재 API 가 받는 5칸으로 옮긴다: 날짜,가맹점,금액,업종코드,사업자번호."""
    out, skipped = [], []
    for n, row in enumerate(rows[header_at + 1:], start=header_at + 2):
        def cell(key):
            i = cols.get(key)
            return '' if i is None or i >= len(row) or row[i] is None else str(row[i]).strip()
        date, merchant, amount = cell('date'), cell('merchant'), cell('amount')
        if not date or not merchant or not amount:
            if any([date, merchant, amount]):
                skipped.append((n, '칸이 비었다', f'{date}|{merchant}|{amount}'))
            continue
        biz = re.sub(r'\D', '', cell('biz'))
        # 금액은 실수로 나오기도 한다(19500.0). 정수로 만든다.
        try:
            amount = str(int(round(float(amount.replace(',', '')))))
        except ValueError:
            skipped.append((n, f'금액을 못 읽음: {amount}', merchant))
            continue
        out.append('%s,"%s",%s,%s,%s' % (date, merchant.replace('"', ''), amount,
                                         cell('industry'), biz if len(biz) == 10 else ''))
    return out, skipped


def seed_numbers():
    if not os.path.exists(SEED):
        return set()
    return {l.split('\t')[0] for l in io.open(SEED, encoding='utf-8')
            if not l.startswith('#') and l.strip()}


def post(base, path, payload):
    req = urllib.request.Request(base + path, method='POST',
                                 data=json.dumps(payload, ensure_ascii=False).encode('utf-8'),
                                 headers={'Content-Type': 'application/json; charset=utf-8'})
    with urllib.request.urlopen(req, timeout=180) as r:
        return json.loads(r.read().decode('utf-8'))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('file')
    ap.add_argument('--name', required=True)
    ap.add_argument('--social7', required=True)
    ap.add_argument('--phone', required=True)
    ap.add_argument('--card-code', type=int, default=None)
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()

    # 실 개인정보다. 저장소가 추적하는 경로에서 읽지 않는다.
    rel = os.path.relpath(os.path.abspath(a.file), ROOT)
    if not rel.startswith('_archive' + os.sep) and not rel.startswith('..'):
        sys.exit(f'명세서는 _archive/ 안에 둔다(저장소 추적 밖): {rel}')

    rows = read_rows(a.file)
    at, cols = find_header(rows)
    if at is None:
        sys.exit('머리글을 못 찾았다 — 날짜·가맹점·금액 칸이 있어야 한다')
    print(f'  머리글 {at + 1}번째 줄 · 칸 {dict(sorted(cols.items()))}')

    lines, skipped = to_csv(rows, at, cols)
    print(f'  읽은 결제 {len(lines)}건' + (f' · 옮기지 못한 줄 {len(skipped)}건' if skipped else ''))
    for n, why, raw in skipped[:5]:
        print(f'     {n}줄  {why}  {raw[:40]}')

    seed = seed_numbers()
    nums = [l.rsplit(',', 1)[1] for l in lines]
    valid = [b for b in nums if len(b) == 10]
    hit = [b for b in valid if b in seed]
    print(f'  사업자번호 있는 결제 {len(valid)}건 · 고유 {len(set(valid))}곳')
    print(f'  **사전 예측** 씨앗과 겹침 {len(hit)}건 · 고유 {len(set(hit))}곳 '
          f'— 적재 뒤 DICT 가 이 수와 맞아야 사슬이 이어진 것이다')

    if a.dry_run:
        print('\n  --dry-run: 아무것도 보내지 않았다. 위 숫자를 확인하고 다시 실행한다.')
        return

    base = os.environ.get('BASE', 'http://localhost:8082')
    body = {'name': a.name, 'social7': a.social7, 'phone': a.phone,
            'cardCode': a.card_code, 'csv': '\n'.join(lines)}
    r = post(base, '/admin/realdata/payments/csv', body)
    print(f"\n  적재 완료 — CI {r['ci'][:12]}… · 카드 {r['cardId']}")
    print(f"     신규 {r['accepted']} · 거부 {r['rejected']} · 채워넣음 {r.get('backfilled', 0)}"
          f" · 사업자번호 실림 {r.get('withBusinessNumber', 0)}")
    for p in (r.get('problems') or [])[:10]:
        print(f"     {p['line']}줄  {p['reason']}")
    if len(r.get('problems') or []) > 10:
        print(f"     … 그 외 {len(r['problems']) - 10}건")


if __name__ == '__main__':
    main()
