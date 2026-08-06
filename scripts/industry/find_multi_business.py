"""한 번호에 **성격이 다른 사업**이 붙은 곳 후보를 찾는다 — `복합사업자-사업자번호.tsv` 의 재료.

  실행:  set -a; . .env; set +a
         .venv/bin/python scripts/industry/find_multi_business.py [명세서.xlsx]
         (기본 `_archive/cardrealdata1.xlsx`)

## 무엇으로 판정하나 — **중분류가 갈리는가**

복합 사업자의 정의는 "상호가 여럿"이 아니라 **"그 번호의 가맹점들이 서로 다른 중분류로 간다"**
이다. 그러면 판정은 *이름을 업종으로 옮기는 일*이고, 그 도구가 이미 있다 — 가맹점명으로 업종을
맞히는 분류기다(`MerchantClassifierService` 와 같은 프롬프트·같은 `toMid` 규칙을 쓴다).

**상호의 어휘로 가르려던 적이 있는데 두 곳에서 틀렸다**(2026-08-05 실측).

    티머니(396-87-03587)       시외버스·택시·카카오택시 — 어휘는 제각각인데 **전부 교통**이다
    포스텍복지회(506-82-61587)  둘 다 `GS25` 로 시작하는데 **편의점과 구내식당**이다

어휘는 증상이고 중분류가 성격이다. 실측 8곳(같은 사업 4 · 복합 4)이 전부 맞았다.

    KTX 23종        → 교통/자동차 하나            → 단일
    티머니 39종      → 교통/자동차 하나            → 단일
    올리브영 3지점    → 미용 하나                  → 단일
    맥도날드 2지점    → 식비 하나                  → 단일
    울릉크루즈       → 교통/자동차 + 편의점/잡화     → **복합**
    롯데백화점       → 쇼핑 + 미용 + 생활          → **복합**
    아이파크몰       → 쇼핑 + 카페/간식            → **복합**
    포스텍복지회      → 편의점/잡화 + 생활          → **복합**

## 그래도 **판정이 아니라 제안**이다

LLM 이 사전 구조를 스스로 바꾸면 "판단은 설명가능한 모델이"(마스터 §4 원칙 1)가 깨진다.
개별 답은 틀릴 수 있다 — `아이파크백)식물학` 을 카페로 봤는데 꽃집일 수도 있다. **갈렸다는
사실**은 그래도 유효하지만, 그걸로 자동 등재하면 오탐이 조용히 쌓인다. 그래서 근거와 함께
후보를 내고, 목록에 적는 것은 사람이 한다.

## 애매하면 **복합 쪽으로** 기운다

두 오류의 무게가 다르다.

    복합인데 단일로 보면   →  **오염**. 하나를 고치면 남까지 바뀌고, 조용히 틀린다
    단일인데 복합으로 보면  →  완화가 꺼져 미분류가 는다. **눈에 보이고** LLM·사용자가 채운다

그래서 답을 못 받은 이름이 섞여 있어도, 받은 것들끼리 갈렸으면 후보로 올린다.

## 호출을 줄이는 법 — 어휘는 **표본 추출에만** 쓴다

차량번호·노선·법인격만 다른 상호는 한 종으로 묶어 **대표 하나만** 묻는다(택시 한 번호에 상호가
74종인 곳이 있다). 묶는 것은 어휘로 하고 **판정은 중분류로 한다** — 어휘를 판정에 쓰지 않는 것이
이 도구의 요점이다.
"""
import io
import json
import os
import re
import sys
import time
import urllib.request
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))
DEFAULT = os.path.join(ROOT, '_archive', 'cardrealdata1.xlsx')
TABLE = os.path.join(ROOT, 'backend', 'src', 'main', 'resources', 'industry-mid.json')

KEY = os.environ.get('GEMINI_API_KEY', '')
MODEL = os.environ.get('GEMINI_MODEL', 'gemini-3.1-flash-lite')
BATCH = 40


def load_rows(path):
    if not path.lower().endswith(('.xlsx', '.xlsm')):
        sys.exit(f'xlsx 만 읽는다: {path}')
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl 이 필요하다:  .venv/bin/pip install openpyxl')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]


def declared(name):
    """이미 선언된 번호 — PG 와 복합. 다시 제안하지 않는다."""
    out = set()
    p = os.path.join(HERE, name)
    if not os.path.exists(p):
        return out
    for line in io.open(p, encoding='utf-8'):
        if line.startswith('#') or not line.strip():
            continue
        num = ''.join(ch for ch in line.split('\t')[0] if ch.isdigit())
        if len(num) == 10:
            out.add(num)
    return out


def sample_key(n):
    """차량번호·노선·법인격처럼 **같은 사업 안에서만 달라지는 부분**을 지운 표본 키.

    <b>판정에는 쓰지 않는다</b> — 물어볼 이름을 줄이는 데만 쓴다.
    """
    x = re.sub(r'[가-힣]{2}\s?\d{2,3}[가-힣]\s?\d{3,4}', '', n)   # 차량번호
    x = re.sub(r'[가-힣]+\s?-\s?[가-힣]+', '', x)                 # 노선(포항-서울)
    x = re.sub(r'\d+', '', x)
    x = re.sub(r'\(?주\)?|주식회사|유한회사|\(유\)|㈜', '', x)
    return re.sub(r'[\s()（）\[\]_·\-]', '', x).upper() or n


def build_prompt(catalog, names):
    lst = '\n'.join(f'{i + 1}. {n}' for i, n in enumerate(names))
    return f"""아래는 한국 카드 명세서에 찍힌 가맹점명입니다. 각 가맹점이 어느 업종인지 고르세요.

업종 목록입니다. 대괄호는 그 업종이 속한 소비 분류이고, 답에는 업종 이름만 쓰세요.

{catalog}
- 가맹점이 무엇을 파는지 알겠다면 **목록에서 가장 가까운 업종**을 고르세요.
- 해외 가맹점도 마찬가지입니다.
- 뜻을 알 수 없는 상호, 사람 이름만 있는 것, 숫자뿐인 것은 빼세요.
- 목록에 있는 이름을 **글자 그대로** 쓰세요.

설명·마크다운 없이 JSON만 출력하세요.
형식: {{"1": "체인화 편의점", "3": "택시 운송업"}}

가맹점:
{lst}
"""


def call(prompt):
    url = f'https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent?key={KEY}'
    body = json.dumps({'contents': [{'parts': [{'text': prompt}]}],
                       'generationConfig': {'temperature': 0}}, ensure_ascii=False)
    req = urllib.request.Request(url, data=body.encode('utf-8'),
                                 headers={'Content-Type': 'application/json'})
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=180) as r:
                d = json.loads(r.read().decode('utf-8'))
            return d['candidates'][0]['content']['parts'][0]['text']
        except Exception as e:            # noqa: BLE001 — 어떤 실패든 재시도 후 포기
            if attempt == 2:
                print(f'   ✗ 호출 실패: {e}', file=sys.stderr)
                return ''
            time.sleep(4)
    return ''


def squash(x):
    return re.sub(r'[\s,()（）·]', '', x or '')


def to_mid(ans, name2mid, mids):
    """자바 `MerchantClassifierService.toMid` 와 같은 규칙 — 정확일치 → 중분류 직답 → 근사일치."""
    a = (ans or '').strip()
    if not a:
        return None
    if a in name2mid:
        return name2mid[a]
    if a in mids:
        return a
    key = squash(a)
    if not key:
        return None
    hit = None
    for nm, mid in name2mid.items():
        n = squash(nm)
        if key not in n and n not in key:
            continue
        if hit is not None and hit != mid:
            return None
        hit = mid
    return hit


def classify(names):
    name2mid = json.load(io.open(TABLE, encoding='utf-8'))['midByIndustryName']
    mids = sorted(set(name2mid.values()))
    by_mid = defaultdict(list)
    for n, m in name2mid.items():
        by_mid[m].append(n)
    catalog = '\n'.join(f'[{m}] ' + ' · '.join(sorted(v)) for m, v in sorted(by_mid.items()))

    out = {}
    for s in range(0, len(names), BATCH):
        chunk = names[s:s + BATCH]
        text = call(build_prompt(catalog, chunk))
        i, j = text.find('{'), text.rfind('}')
        try:
            raw = json.loads(text[i:j + 1]) if i >= 0 and j > i else {}
        except ValueError:
            raw = {}
        for k, v in raw.items():
            try:
                idx = int(str(k).strip())
            except ValueError:
                continue
            if 1 <= idx <= len(chunk):
                out[chunk[idx - 1]] = to_mid(str(v).strip(), name2mid, mids)
        print(f'      … {min(s + BATCH, len(names))}/{len(names)}종')
        time.sleep(1)
    return out


def main():
    if not KEY:
        sys.exit('GEMINI_API_KEY 가 없다.  set -a; . .env; set +a  로 넣고 다시 실행한다')
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT
    rows = load_rows(path)
    skip = declared('pg-사업자번호.tsv') | declared('복합사업자-사업자번호.tsv')

    # 번호 → 표본키 → [대표 상호, 건수, 금액]
    by_biz = defaultdict(lambda: defaultdict(lambda: [None, 0, 0.0]))
    for r in rows[4:]:
        if not r or not r[0] or len(r) < 7 or r[6] is None:
            continue
        try:
            amount = float(r[6])
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue
        biz = re.sub(r'\D', '', str(r[5] or ''))
        if len(biz) != 10 or biz in skip:
            continue
        name = str(r[4]).strip()
        cell = by_biz[biz][sample_key(name)]
        if cell[0] is None:
            cell[0] = name
        cell[1] += 1
        cell[2] += amount

    targets = {b: g for b, g in by_biz.items() if len(g) >= 2}
    reps = sorted({c[0] for g in targets.values() for c in g.values()})
    print(f'   명세서: {os.path.relpath(path, ROOT)}')
    print(f'   이미 선언된 번호 {len(skip)}곳 제외 · 표본이 둘 이상인 번호 {len(targets)}곳')
    print(f'   가맹점 {len(reps)}종을 분류기에 묻는다 (모델 {MODEL})')

    mid_of = classify(reps)
    print()

    found = 0
    for biz, group in sorted(targets.items(), key=lambda x: -sum(c[2] for c in x[1].values())):
        got = {c[0]: mid_of.get(c[0]) for c in group.values()}
        distinct = {m for m in got.values() if m}
        if len(distinct) < 2:
            continue
        found += 1
        total = sum(c[2] for c in group.values())
        print(f'   {biz[:3]}-{biz[3:5]}-{biz[5:]}  중분류 {len(distinct)}종 · {total:,.0f}원'
              f'   {sorted(distinct)}')
        for c in sorted(group.values(), key=lambda x: -x[2])[:6]:
            print(f'        {c[1]:>3}건 {c[2]:>10,.0f}원  {c[0][:36]:<38} '
                  f'{got.get(c[0]) or "(답 없음)"}')
        print()

    if found == 0:
        print('   새 후보 없음 — 상호가 여럿인 번호들이 전부 한 중분류로 모였다.')
    else:
        print(f'   후보 {found}곳. 근거를 보고 맞으면 scripts/industry/복합사업자-사업자번호.tsv 에 적는다.')
        print('   **애매하면 복합 쪽으로 기운다** — 오염은 조용히 틀리고, 미분류는 눈에 보인다.')


if __name__ == '__main__':
    main()
